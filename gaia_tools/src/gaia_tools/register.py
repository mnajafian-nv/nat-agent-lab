"""Custom GAIA benchmark tools for NeMo Agent Toolkit.

Seven tools that unlock the majority of GAIA questions:
  - read_file: Extract text from local files (PDF, CSV, Excel, JSON, HTML, text)
  - fetch_url: Download and extract content from URLs
  - python_executor: Sandboxed Python code execution for calculations
  - describe_image / describe_image_alt: Describe images using vision models
  - transcribe_audio: Transcribe audio files using NVIDIA Parakeet ASR
  - get_youtube_transcript: Fetch YouTube video captions (no API key needed)
  - solve_chess: Composite tool for chess position analysis (vision + python-chess)
"""

import asyncio
import base64
import logging
import os
import re
import subprocess
import tempfile
from pathlib import Path
from typing import Optional

from pydantic import Field

from nat.builder.builder import Builder
from nat.builder.function_info import FunctionInfo
from nat.cli.register_workflow import register_function
from nat.data_models.function import FunctionBaseConfig

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Tool 1: read_file
# ---------------------------------------------------------------------------

class ReadFileToolConfig(FunctionBaseConfig, name="read_file"):
    """Read and extract text from a local file."""
    max_chars: int = Field(default=50000, description="Maximum characters to return from file content")


def _extract_pdf(file_path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(file_path)
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"[Page {i + 1}]\n{text}")
    return "\n\n".join(pages) if pages else "PDF contained no extractable text."


def _extract_csv(file_path: str) -> str:
    import csv
    rows = []
    with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            rows.append(",".join(row))
            if i >= 500:
                rows.append(f"... (truncated, {i + 1}+ rows total)")
                break
    return "\n".join(rows)


def _extract_excel(file_path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl not installed. Cannot read Excel files."
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(",".join(str(c) if c is not None else "" for c in row))
            if i >= 500:
                rows.append("... (truncated)")
                break
        sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(sheets)


def _extract_html(file_path: str) -> str:
    from bs4 import BeautifulSoup
    with open(file_path, encoding="utf-8", errors="replace") as f:
        soup = BeautifulSoup(f.read(), "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    return soup.get_text(separator="\n", strip=True)


def _extract_json(file_path: str) -> str:
    import json
    with open(file_path, encoding="utf-8") as f:
        data = json.load(f)
    return json.dumps(data, indent=2, ensure_ascii=False)


def _extract_text(file_path: str) -> str:
    with open(file_path, encoding="utf-8", errors="replace") as f:
        return f.read()


def _read_local_file(file_path: str, max_chars: int) -> str:
    path = Path(file_path)
    if not path.exists():
        return f"Error: File not found: {file_path}"

    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            content = _extract_pdf(file_path)
        elif suffix == ".csv":
            content = _extract_csv(file_path)
        elif suffix in (".xlsx", ".xls"):
            content = _extract_excel(file_path)
        elif suffix in (".html", ".htm"):
            content = _extract_html(file_path)
        elif suffix == ".json":
            content = _extract_json(file_path)
        else:
            content = _extract_text(file_path)
    except Exception as e:
        return f"Error reading {file_path}: {type(e).__name__}: {e}"

    if len(content) > max_chars:
        content = content[:max_chars] + f"\n... (truncated at {max_chars} chars)"
    return content


@register_function(config_type=ReadFileToolConfig)
async def read_file(config: ReadFileToolConfig, builder: Builder):

    async def _read_file(file_path: str) -> str:
        """Read and extract text from a local file. Supports PDF, CSV, Excel (.xlsx),
        JSON, HTML, and plain text files. Use this when a question references an
        attached file or when you need to examine file contents.

        Args:
            file_path: Absolute or relative path to the file to read.

        Returns:
            The extracted text content of the file.
        """
        return await asyncio.to_thread(_read_local_file, file_path, config.max_chars)

    yield FunctionInfo.from_fn(_read_file, description=_read_file.__doc__)


# ---------------------------------------------------------------------------
# Tool 2: fetch_url
# ---------------------------------------------------------------------------

class FetchURLToolConfig(FunctionBaseConfig, name="fetch_url"):
    """Fetch and extract text content from a URL."""
    max_chars: int = Field(default=50000, description="Maximum characters to return")
    timeout: int = Field(default=30, description="HTTP request timeout in seconds")


def _fetch_and_extract(url: str, max_chars: int, timeout: int) -> str:
    import requests
    from bs4 import BeautifulSoup

    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; GAIA-Agent/1.0)",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        return f"Error fetching {url}: {type(e).__name__}: {e}"

    content_type = resp.headers.get("Content-Type", "").lower()

    if "application/pdf" in content_type or url.lower().endswith(".pdf"):
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        try:
            content = _extract_pdf(tmp_path)
        finally:
            os.unlink(tmp_path)
    elif "text/csv" in content_type or url.lower().endswith(".csv"):
        content = resp.text
    elif "spreadsheet" in content_type or url.lower().endswith((".xlsx", ".xls")):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        try:
            content = _extract_excel(tmp_path)
        finally:
            os.unlink(tmp_path)
    elif "text/html" in content_type:
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        content = soup.get_text(separator="\n", strip=True)
    elif "application/json" in content_type:
        import json
        try:
            data = resp.json()
            content = json.dumps(data, indent=2, ensure_ascii=False)
        except Exception:
            content = resp.text
    else:
        content = resp.text

    if len(content) > max_chars:
        content = content[:max_chars] + f"\n... (truncated at {max_chars} chars)"
    return content


@register_function(config_type=FetchURLToolConfig)
async def fetch_url(config: FetchURLToolConfig, builder: Builder):

    async def _fetch_url(url: str) -> str:
        """Fetch and extract text content from a URL. Handles HTML pages, PDFs, CSV
        files, Excel spreadsheets, and JSON. Use this to read web pages, download
        and parse documents from the internet, or access data at a specific URL.

        Args:
            url: The full URL to fetch (must start with http:// or https://).

        Returns:
            The extracted text content from the URL.
        """
        return await asyncio.to_thread(_fetch_and_extract, url, config.max_chars, config.timeout)

    yield FunctionInfo.from_fn(_fetch_url, description=_fetch_url.__doc__)


# ---------------------------------------------------------------------------
# Tool 3: python_executor
# ---------------------------------------------------------------------------

class PythonExecutorToolConfig(FunctionBaseConfig, name="python_executor"):
    """Execute Python code and return the output."""
    timeout: int = Field(default=30, description="Execution timeout in seconds")


@register_function(config_type=PythonExecutorToolConfig)
async def python_executor(config: PythonExecutorToolConfig, builder: Builder):

    async def _run_python(code: str) -> str:
        """Execute Python code and return the printed output. Use this for mathematical
        calculations, data processing, string manipulation, date arithmetic, or any
        computation that needs precise results. The code runs in a fresh Python process.

        Tips:
        - Use print() to output results; the last print() value is returned.
        - Standard library modules are available (math, datetime, json, csv, re, etc.).
        - For complex math, use the math or statistics modules.
        - Keep code concise and focused on the calculation needed.

        Args:
            code: Python code to execute. Must use print() to output results.

        Returns:
            The stdout output from executing the code, or an error message.
        """
        def _exec():
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".py", delete=False, encoding="utf-8"
            ) as tmp:
                tmp.write(code)
                tmp_path = tmp.name
            try:
                result = subprocess.run(
                    ["python3", tmp_path],
                    capture_output=True,
                    text=True,
                    timeout=config.timeout,
                    env={**os.environ, "PYTHONDONTWRITEBYTECODE": "1"},
                )
                output = result.stdout.strip()
                if result.returncode != 0:
                    err = result.stderr.strip()
                    if output:
                        return f"{output}\n\nError (exit code {result.returncode}):\n{err}"
                    return f"Error (exit code {result.returncode}):\n{err}"
                return output if output else "(no output)"
            except subprocess.TimeoutExpired:
                return f"Error: Code execution timed out after {config.timeout} seconds."
            except Exception as e:
                return f"Error executing code: {type(e).__name__}: {e}"
            finally:
                os.unlink(tmp_path)

        return await asyncio.to_thread(_exec)

    yield FunctionInfo.from_fn(_run_python, description=_run_python.__doc__)


# ---------------------------------------------------------------------------
# Tool 4: describe_image
# ---------------------------------------------------------------------------

class DescribeImageToolConfig(FunctionBaseConfig, name="describe_image"):
    """Describe an image using a vision-capable LLM."""
    model: str = Field(default="meta/llama-3.2-90b-vision-instruct", description="Vision model to use")
    api_key_env: str = Field(default="NGC_API_KEY", description="Env var holding the API key")
    base_url: Optional[str] = Field(default="https://integrate.api.nvidia.com/v1", description="Base URL for OpenAI-compatible vision API")
    max_tokens: int = Field(default=1024, description="Max tokens for the description response")


def _describe_image_sync(
    image_source: str,
    question: str,
    model: str,
    api_key: str,
    base_url: Optional[str],
    max_tokens: int,
) -> str:
    from openai import OpenAI

    client_kwargs = {"api_key": api_key, "timeout": 45, "max_retries": 1}
    if base_url:
        client_kwargs["base_url"] = base_url
    client = OpenAI(**client_kwargs)

    content_parts = []
    if question:
        content_parts.append({"type": "text", "text": question})

    if image_source.startswith(("http://", "https://")):
        content_parts.append({
            "type": "image_url",
            "image_url": {"url": image_source},
        })
    else:
        path = Path(image_source)
        if not path.exists():
            return f"Error: Image file not found: {image_source}"
        suffix = path.suffix.lower().lstrip(".")
        mime_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp"}
        mime = mime_map.get(suffix, "image/png")
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        content_parts.append({
            "type": "image_url",
            "image_url": {"url": f"data:{mime};base64,{b64}"},
        })

    if not question:
        content_parts.insert(0, {
            "type": "text",
            "text": "Describe this image in detail. Include all text, numbers, labels, and visual elements.",
        })

    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": content_parts}],
            max_tokens=max_tokens,
            temperature=0,
        )
        content = resp.choices[0].message.content
        return content.strip() if content else "Error: vision model returned empty response"
    except Exception as e:
        return f"Error describing image: {type(e).__name__}: {e}"


@register_function(config_type=DescribeImageToolConfig)
async def describe_image(config: DescribeImageToolConfig, builder: Builder):

    api_key = os.environ.get(config.api_key_env, "")
    if not api_key:
        logger.warning("describe_image: %s not set; tool will fail at runtime", config.api_key_env)

    async def _describe_image(image_source: str, question: str = "") -> str:
        """Describe or analyze an image using a vision model. Use this when a GAIA
        question includes an image attachment (chart, screenshot, diagram, photo,
        etc.) that needs visual understanding. Do NOT use for chess positions;
        use solve_chess instead.

        Args:
            image_source: Path to a local image file OR an image URL (http/https).
            question: Optional specific question about the image. If empty, returns
                a general detailed description.

        Returns:
            A text description or answer based on the image content.
        """
        return await asyncio.to_thread(
            _describe_image_sync,
            image_source, question, config.model,
            api_key, config.base_url, config.max_tokens,
        )

    yield FunctionInfo.from_fn(_describe_image, description=_describe_image.__doc__)


# ---------------------------------------------------------------------------
# Tool 5: transcribe_audio
# ---------------------------------------------------------------------------

class TranscribeAudioToolConfig(FunctionBaseConfig, name="transcribe_audio"):
    """Transcribe audio files to text using NVIDIA Parakeet ASR."""
    api_key_env: str = Field(default="NGC_API_KEY", description="Env var holding the NVIDIA API key")
    function_id: str = Field(
        default="1598d209-5e27-4d3c-8079-4751568b1081",
        description="NVCF function ID for parakeet-ctc-1.1b-asr",
    )
    grpc_server: str = Field(default="grpc.nvcf.nvidia.com:443", description="gRPC server address")
    language_code: str = Field(default="en-US", description="Language code for transcription")


def _convert_to_wav(file_path: str) -> str:
    """Convert audio to 16-bit mono WAV using ffmpeg. Returns path to WAV file."""
    wav_path = file_path.rsplit(".", 1)[0] + "_converted.wav"
    try:
        subprocess.run(
            ["ffmpeg", "-y", "-i", file_path, "-ar", "16000", "-ac", "1",
             "-sample_fmt", "s16", wav_path],
            capture_output=True, text=True, timeout=30,
        )
        if Path(wav_path).exists():
            return wav_path
    except Exception:
        pass
    return file_path


def _transcribe_with_riva(
    file_path: str, api_key: str, function_id: str,
    grpc_server: str, language_code: str,
) -> str:
    try:
        import riva.client
    except ImportError:
        return "Error: nvidia-riva-client not installed. Run: pip install nvidia-riva-client"

    path = Path(file_path)
    if not path.exists():
        return f"Error: Audio file not found: {file_path}"

    if path.suffix.lower() in (".mp3", ".m4a", ".ogg", ".flac", ".aac"):
        file_path = _convert_to_wav(file_path)

    with open(file_path, "rb") as f:
        audio_data = f.read()

    try:
        metadata = [
            ("function-id", function_id),
            ("authorization", f"Bearer {api_key}"),
        ]
        auth = riva.client.Auth(
            use_ssl=True,
            uri=grpc_server,
            metadata_args=metadata,
        )
        asr_service = riva.client.ASRService(auth)

        config = riva.client.RecognitionConfig(
            language_code=language_code,
            max_alternatives=1,
            enable_automatic_punctuation=True,
            audio_channel_count=1,
        )

        response = asr_service.offline_recognize(audio_data, config)

        transcript_parts = []
        for result in response.results:
            if result.alternatives:
                transcript_parts.append(result.alternatives[0].transcript)

        transcript = " ".join(transcript_parts).strip()
        return transcript if transcript else "No speech detected in audio."
    except Exception as e:
        return f"Error transcribing audio: {type(e).__name__}: {e}"


@register_function(config_type=TranscribeAudioToolConfig)
async def transcribe_audio(config: TranscribeAudioToolConfig, builder: Builder):

    api_key = os.environ.get(config.api_key_env, "")
    if not api_key:
        logger.warning("transcribe_audio: %s not set; tool will fail at runtime", config.api_key_env)

    async def _transcribe_audio(file_path: str) -> str:
        """Transcribe an audio file to text. Use this when a GAIA question includes
        an audio attachment (mp3, wav, m4a, ogg, flac) that contains spoken content
        you need to understand. The tool converts speech to text.

        Args:
            file_path: Path to the audio file to transcribe.

        Returns:
            The transcribed text from the audio file.
        """
        return await asyncio.to_thread(
            _transcribe_with_riva,
            file_path, api_key, config.function_id,
            config.grpc_server, config.language_code,
        )

    yield FunctionInfo.from_fn(_transcribe_audio, description=_transcribe_audio.__doc__)


# ---------------------------------------------------------------------------
# Tool 6: get_youtube_transcript
# ---------------------------------------------------------------------------

class YouTubeTranscriptToolConfig(FunctionBaseConfig, name="get_youtube_transcript"):
    """Fetch the transcript/captions of a YouTube video."""
    languages: list[str] = Field(default=["en"], description="Preferred transcript languages")


def _extract_video_id(video_id_or_url: str) -> str:
    """Extract a bare video ID from a full YouTube URL or return as-is."""
    # https://www.youtube.com/watch?v=XXXXX or https://youtu.be/XXXXX
    m = re.search(r"(?:v=|youtu\.be/)([\w-]{11})", video_id_or_url)
    if m:
        return m.group(1)
    # Already a bare ID (11 chars of word chars / hyphens)
    stripped = video_id_or_url.strip()
    if re.fullmatch(r"[\w-]{11}", stripped):
        return stripped
    return stripped


def _fetch_transcript(video_id: str, languages: list[str]) -> str:
    try:
        from youtube_transcript_api import YouTubeTranscriptApi
    except ImportError:
        return "Error: youtube-transcript-api not installed. Run: pip install youtube-transcript-api"

    video_id = _extract_video_id(video_id)
    try:
        ytt = YouTubeTranscriptApi()
        transcript = ytt.fetch(video_id, languages=languages)
        lines = [snippet.text for snippet in transcript]
        return " ".join(lines)
    except Exception as e:
        return f"Error fetching transcript for video {video_id}: {type(e).__name__}: {e}"


@register_function(config_type=YouTubeTranscriptToolConfig)
async def get_youtube_transcript(config: YouTubeTranscriptToolConfig, builder: Builder):

    async def _get_youtube_transcript(video_id: str) -> str:
        """Fetch the transcript (captions/subtitles) of a YouTube video. Use this
        whenever a question references a YouTube video URL. Extract the video ID from
        the URL: for https://www.youtube.com/watch?v=XXXXX the ID is XXXXX.

        This returns the actual spoken words from the video, which is far more
        reliable than searching the web for descriptions of the video content.

        Args:
            video_id: The YouTube video ID (the part after v= in the URL).

        Returns:
            The full transcript text of the video.
        """
        return await asyncio.to_thread(_fetch_transcript, video_id, config.languages)

    yield FunctionInfo.from_fn(_get_youtube_transcript, description=_get_youtube_transcript.__doc__)


# ---------------------------------------------------------------------------
# Tool 7: solve_chess
# ---------------------------------------------------------------------------

class SolveChessToolConfig(FunctionBaseConfig, name="solve_chess"):
    """Solve chess positions from board images using vision models + python-chess."""
    primary_model: str = Field(default="qwen/qwen3.5-397b-a17b", description="Primary vision model")
    secondary_model: str = Field(default="mistralai/mistral-large-3-675b-instruct-2512", description="Secondary vision model")
    api_key_env: str = Field(default="NGC_API_KEY", description="Env var holding the API key")
    base_url: str = Field(default="https://integrate.api.nvidia.com/v1", description="Vision API base URL")
    max_tokens: int = Field(default=2048, description="Max tokens for vision responses")


def _flip_fen_perspective(fen_placement: str) -> str:
    """Rotate a FEN piece-placement string 180 degrees.

    Handles the case where the vision model read a board shown from Black's
    perspective but assumed White's perspective (row 8 top, column a left).
    """
    rows = fen_placement.split("/")
    flipped = []
    for row in reversed(rows):
        expanded = []
        for ch in row:
            if ch.isdigit():
                expanded.extend(["."] * int(ch))
            else:
                expanded.append(ch)
        expanded.reverse()
        compressed: list[str] = []
        empty = 0
        for ch in expanded:
            if ch == ".":
                empty += 1
            else:
                if empty:
                    compressed.append(str(empty))
                    empty = 0
                compressed.append(ch)
        if empty:
            compressed.append(str(empty))
        flipped.append("".join(compressed))
    return "/".join(flipped)


def _load_image_content(image_source: str):
    """Build the image content part for an OpenAI vision request. Returns (content_part, error)."""
    if image_source.startswith(("http://", "https://")):
        return {"type": "image_url", "image_url": {"url": image_source}}, None
    path = Path(image_source)
    if not path.exists():
        return None, f"ERROR: Image file not found: {image_source}"
    suffix = path.suffix.lower().lstrip(".")
    mime_map = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
                "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp"}
    mime = mime_map.get(suffix, "image/png")
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}}, None


def _call_vision_with_client(client, image_content, prompt: str, model: str, max_tokens: int, timeout: int = 30) -> str:
    """Call a vision model with a pre-loaded image. Returns response text or error string."""
    content_parts = [{"type": "text", "text": prompt}, image_content]
    try:
        resp = client.chat.completions.create(
            model=model, messages=[{"role": "user", "content": content_parts}],
            max_tokens=max_tokens, timeout=timeout, temperature=0,
        )
        if not resp or not resp.choices:
            return "ERROR: No response from vision model"
        content = resp.choices[0].message.content
        if not content:
            return "ERROR: Empty response from vision model"
        return content.strip()
    except Exception as e:
        return f"ERROR: {type(e).__name__}: {e}"


def _solve_chess_sync(image_source: str, question: str, config: SolveChessToolConfig, api_key: str) -> str:
    import chess
    import shutil
    from openai import OpenAI

    question_lower = question.lower() if question else ""

    black_re = r"black(?:'s| to) (?:turn|move|play)|move for black|best move for black"
    white_re = r"white(?:'s| to) (?:turn|move|play)|move for white|best move for white"
    has_black = bool(re.search(black_re, question_lower))
    has_white = bool(re.search(white_re, question_lower))
    if has_black and not has_white:
        turns_to_try = [chess.BLACK]
    elif has_white and not has_black:
        turns_to_try = [chess.WHITE]
    elif "black" in question_lower:
        turns_to_try = [chess.BLACK, chess.WHITE]
    elif "white" in question_lower:
        turns_to_try = [chess.WHITE, chess.BLACK]
    else:
        turns_to_try = [chess.WHITE, chess.BLACK]

    logger.info("solve_chess: turns_to_try=%s, question=%r", turns_to_try, question[:120])

    FEN_PROMPT = (
        "Convert this chess board image to FEN (Forsyth-Edwards Notation). "
        "Output ONLY the FEN string for piece placement (the part before the first space), nothing else. "
        "Use standard FEN: uppercase=White, lowercase=Black, /=row separator, digits=empty squares. "
        "Row 8 (top) first, row 1 (bottom) last."
    )

    GRID_PROMPT = (
        "Analyze this chess board image carefully. White pieces are at the bottom (ranks 1-2), "
        "Black pieces are at the top (ranks 7-8). Rank 8 is the TOP row.\n\n"
        "List every square from rank 8 (top) to rank 1 (bottom), left (a-file) to right (h-file).\n"
        "Format: square=code where codes are: wp wr wn wb wq wk bp br bn bb bq bk empty\n\n"
        "Example for 2 ranks:\n"
        "Rank 8: a8=br, b8=empty, c8=bb, d8=bq, e8=bk, f8=empty, g8=empty, h8=br\n"
        "Rank 7: a7=bp, b7=bp, c7=empty, d7=empty, e7=bp, f7=bp, g7=bp, h7=bp\n\n"
        "Now do all 8 ranks for this board:\n"
        "Rank 8: a8=?, b8=?, c8=?, d8=?, e8=?, f8=?, g8=?, h8=?\n"
        "Rank 7: a7=?, b7=?, c7=?, d7=?, e7=?, f7=?, g7=?, h7=?\n"
        "Rank 6: a6=?, b6=?, c6=?, d6=?, e6=?, f6=?, g6=?, h6=?\n"
        "Rank 5: a5=?, b5=?, c5=?, d5=?, e5=?, f5=?, g5=?, h5=?\n"
        "Rank 4: a4=?, b4=?, c4=?, d4=?, e4=?, f4=?, g4=?, h4=?\n"
        "Rank 3: a3=?, b3=?, c3=?, d3=?, e3=?, f3=?, g3=?, h3=?\n"
        "Rank 2: a2=?, b2=?, c2=?, d2=?, e2=?, f2=?, g2=?, h2=?\n"
        "Rank 1: a1=?, b1=?, c1=?, d1=?, e1=?, f1=?, g1=?, h1=?\n\n"
        "Fill in every square. Do not skip any."
    )

    GRID_CODE_TO_SYMBOL = {
        "wp": "P", "wr": "R", "wn": "N", "wb": "B", "wq": "Q", "wk": "K",
        "bp": "p", "br": "r", "bn": "n", "bb": "b", "bq": "q", "bk": "k",
    }

    image_content, img_err = _load_image_content(image_source)
    if img_err:
        return img_err
    client = OpenAI(api_key=api_key, base_url=config.base_url, max_retries=1)

    def call_vision(prompt: str, model: str) -> str:
        return _call_vision_with_client(client, image_content, prompt, model, config.max_tokens, timeout=30)

    def try_fen_boards(fen_str: str, turns):
        """Try all (orientation, turn, castling) combos. Return list of valid boards."""
        fen_str = fen_str.strip().split()[0]
        if len(fen_str.split("/")) != 8:
            return []
        fen_variants = [fen_str]
        flipped = _flip_fen_perspective(fen_str)
        if flipped != fen_str:
            fen_variants.append(flipped)
        candidates = []
        for fen in fen_variants:
            for turn_color in turns:
                turn_ch = "b" if turn_color == chess.BLACK else "w"
                for castling in ["-", "KQkq"]:
                    try:
                        full_fen = f"{fen} {turn_ch} {castling} - 0 1"
                        board = chess.Board(full_fen)
                        if board.king(chess.WHITE) is None or board.king(chess.BLACK) is None:
                            continue
                        if not board.is_valid():
                            continue
                        candidates.append(board)
                    except Exception:
                        continue
        return candidates

    def parse_grid(text: str):
        """Parse structured grid response into (square, piece_code) pairs, then build board."""
        pattern = r'([a-h][1-8])\s*[=:\-]\s*(w[prnbqk]|b[prnbqk]|empty|none|[_\.\-]{1,2})'
        matches = re.findall(pattern, text, re.IGNORECASE)
        logger.info("solve_chess: parse_grid matched %d squares", len(matches))
        if len(matches) < 4:
            return []
        candidates = []
        for turn_color in turns_to_try:
            board = chess.Board(None)
            board.turn = turn_color
            try:
                for sq_str, code in matches:
                    code_l = code.lower()
                    if code_l in ("empty", "none", ".", "-", "_", "--", "__"):
                        continue
                    sym = GRID_CODE_TO_SYMBOL.get(code_l)
                    if sym:
                        board.set_piece_at(chess.parse_square(sq_str.lower()), chess.Piece.from_symbol(sym))
                if board.king(chess.WHITE) is None or board.king(chess.BLACK) is None:
                    continue
                if not board.is_valid():
                    logger.info("solve_chess: parse_grid board invalid (turn=%s), skipping", board.turn)
                    continue
                candidates.append(board)
            except Exception:
                continue
        return candidates

    def parse_pieces(text: str):
        """Extract (color, piece_type, square) from natural language."""
        pattern = r'(White|Black)\s+(King|Queen|Rook|Bishop|Knight|Pawn)\s*(?:on|at|:)?\s*([a-h][1-8])'
        return re.findall(pattern, text, re.IGNORECASE)

    PIECE_SYMBOL = {
        ('white', 'king'): 'K', ('white', 'queen'): 'Q', ('white', 'rook'): 'R',
        ('white', 'bishop'): 'B', ('white', 'knight'): 'N', ('white', 'pawn'): 'P',
        ('black', 'king'): 'k', ('black', 'queen'): 'q', ('black', 'rook'): 'r',
        ('black', 'bishop'): 'b', ('black', 'knight'): 'n', ('black', 'pawn'): 'p',
    }

    def build_piece_boards(pieces, turns):
        """Build boards for all requested turns from parsed pieces."""
        candidates = []
        for turn_color in turns:
            board = chess.Board(None)
            board.turn = turn_color
            try:
                for color, ptype, sq in pieces:
                    sym = PIECE_SYMBOL.get((color.lower(), ptype.lower()))
                    if sym:
                        board.set_piece_at(chess.parse_square(sq.lower()), chess.Piece.from_symbol(sym))
                if board.king(chess.WHITE) is None or board.king(chess.BLACK) is None:
                    continue
                candidates.append(board)
            except Exception:
                continue
        return candidates

    # ---- Engine-based best move (Stockfish) ----
    stockfish_path = shutil.which("stockfish")
    if not stockfish_path:
        for fallback in ["/usr/games/stockfish", "/usr/bin/stockfish", "/usr/local/bin/stockfish"]:
            if Path(fallback).is_file():
                stockfish_path = fallback
                break
    has_stockfish = stockfish_path is not None

    def engine_best_move(board: chess.Board) -> str:
        """Use Stockfish to find the best move. Returns SAN string."""
        import chess.engine
        engine = None
        try:
            engine = chess.engine.SimpleEngine.popen_uci(stockfish_path)
            result = engine.play(board, chess.engine.Limit(time=2.0))
            info = engine.analyse(board, chess.engine.Limit(time=1.0))
            san = board.san(result.move)
            score = info.get("score")
            score_str = str(score.relative) if score else "?"
            return f"{san} (eval: {score_str})"
        except Exception as e:
            logger.warning("solve_chess: Stockfish failed: %s", e)
            return ""
        finally:
            if engine:
                try:
                    engine.quit()
                except Exception:
                    pass

    def find_forced_mate(board: chess.Board, max_depth: int = 2):
        """Return SAN of first move leading to forced checkmate within max_depth plies."""
        for move in board.legal_moves:
            san = board.san(move)
            board.push(move)
            if board.is_checkmate():
                board.pop()
                return san
            board.pop()
        if max_depth < 3:
            return None
        for move in board.legal_moves:
            san = board.san(move)
            board.push(move)
            if board.is_game_over():
                board.pop()
                continue
            forced = True
            for opp_move in board.legal_moves:
                board.push(opp_move)
                has_mate_reply = False
                for reply in board.legal_moves:
                    board.push(reply)
                    if board.is_checkmate():
                        has_mate_reply = True
                        board.pop()
                        break
                    board.pop()
                board.pop()
                if not has_mate_reply:
                    forced = False
                    break
            board.pop()
            if forced:
                return san
        return None

    BEST_MOVE_FOUND = object()

    def solve_board(board: chess.Board) -> tuple:
        """Find the best move: checkmate first, then Stockfish, then legal move list."""
        mate = find_forced_mate(board, max_depth=3)
        if mate:
            return (BEST_MOVE_FOUND, mate)
        if has_stockfish:
            best = engine_best_move(board)
            if best:
                move_san = best.split(" (")[0]
                logger.info("solve_chess: Stockfish best move: %s", best)
                return (BEST_MOVE_FOUND, move_san)
        legal = [board.san(m) for m in board.legal_moves]
        checks = [board.san(m) for m in board.legal_moves
                   if board.gives_check(m)]
        parts = [f"No forced checkmate found. {len(legal)} legal moves."]
        if checks:
            parts.append(f"Checking moves: {', '.join(checks[:10])}")
        parts.append(f"All legal moves: {', '.join(legal[:30])}")
        parts.append(f"Board:\n{board}")
        return (None, "\n".join(parts))

    def check_boards(boards, label):
        """Check a list of candidate boards for best move. Returns result or None."""
        for board in boards:
            sentinel, result = solve_board(board)
            if sentinel is BEST_MOVE_FOUND:
                logger.info("solve_chess: Best move found via %s: %s", label, result)
                return result
            boards_found.append((board, label))
        return None

    primary = config.primary_model
    secondary = config.secondary_model
    boards_found = []
    debug_info = []

    if has_stockfish:
        logger.info("solve_chess: Stockfish available at %s", stockfish_path)
    else:
        logger.info("solve_chess: Stockfish NOT found; will use checkmate search only")

    def try_model_strategies(model, label_prefix):
        """Run Grid then FEN for a single model. Returns answer string or None."""
        resp = call_vision(GRID_PROMPT, model)
        if not resp.startswith("ERROR"):
            boards = parse_grid(resp)
            if boards:
                logger.info("solve_chess: %s Grid produced %d board(s)", label_prefix, len(boards))
                result = check_boards(boards, f"{label_prefix} Grid")
                if result:
                    return result
                debug_info.append(f"{label_prefix} Grid: {len(boards)} board(s), no winning move")
            else:
                debug_info.append(f"{label_prefix} Grid: parse failed")
                logger.info("solve_chess: %s Grid parse failed", label_prefix)
        else:
            debug_info.append(f"{label_prefix} Grid: {resp[:120]}")
            logger.warning("solve_chess: %s Grid error: %s", label_prefix, resp[:200])

        resp = call_vision(FEN_PROMPT, model)
        if not resp.startswith("ERROR"):
            for token in resp.replace("\n", " ").split():
                token = token.strip("`'\" ")
                boards = try_fen_boards(token, turns_to_try)
                if boards:
                    logger.info("solve_chess: %s FEN produced %d board(s)", label_prefix, len(boards))
                    result = check_boards(boards, f"{label_prefix} FEN")
                    if result:
                        return result
                    debug_info.append(f"{label_prefix} FEN: {len(boards)} board(s), no winning move")
                    break
            else:
                debug_info.append(f"{label_prefix} FEN: no valid FEN parsed")
                logger.info("solve_chess: %s FEN: no valid FEN", label_prefix)
        else:
            debug_info.append(f"{label_prefix} FEN: {resp[:120]}")
            logger.warning("solve_chess: %s FEN error: %s", label_prefix, resp[:200])
        return None

    # ---- Phase 1: primary model (2 API calls max) ----
    result = try_model_strategies(primary, "Primary")
    if result:
        return result

    # If primary gave us boards, try Stockfish fallback before calling secondary
    if boards_found and has_stockfish:
        board, source = boards_found[0]
        best = engine_best_move(board)
        if best:
            move_san = best.split(" (")[0]
            logger.info("solve_chess: Stockfish on primary board (%s): %s", source, best)
            return move_san

    # ---- Phase 2: secondary model (only if primary produced no boards) ----
    if not boards_found:
        logger.info("solve_chess: Primary produced no boards, trying secondary model")
        result = try_model_strategies(secondary, "Secondary")
        if result:
            return result

    # ---- Fallback: best available info ----
    if boards_found:
        board, source = boards_found[0]
        _sentinel, fallback_result = solve_board(board)
        logger.info("solve_chess: Final fallback from %s", source)
        return fallback_result

    # ---- Last resort: ask vision model directly for the move ----
    logger.warning("solve_chess: No valid board from any strategy; asking vision for move directly")
    direct_prompt = (
        f"Look at this chess board image. {question}\n"
        "Answer with ONLY the move in standard algebraic notation (e.g., Qh7#, Nf3, Bxc6). "
        "Nothing else."
    )
    move_re = re.compile(r'\b([KQRBN]?[a-h]?[1-8]?x?[a-h][1-8](?:=[QRBN])?[+#]?)\b')
    for model in [primary, secondary]:
        resp = call_vision(direct_prompt, model)
        if not resp.startswith("ERROR"):
            move_match = move_re.search(resp)
            if move_match:
                logger.info("solve_chess: Direct vision move from %s: %s", model, move_match.group(0))
                return move_match.group(0)

    logger.warning("solve_chess: All strategies exhausted. Debug: %s", "; ".join(debug_info))
    return "Unable to determine the best move from this position."


@register_function(config_type=SolveChessToolConfig)
async def solve_chess(config: SolveChessToolConfig, builder: Builder):

    api_key = os.environ.get(config.api_key_env, "")
    if not api_key:
        logger.warning("solve_chess: %s not set; tool will fail at runtime", config.api_key_env)

    async def _solve_chess(image_source: str, question: str = "") -> str:
        """Solve a chess position from a board image. Identifies pieces using
        vision models and finds checkmate or best moves using python-chess.

        Use this INSTEAD of describe_image for chess positions. This tool handles
        the entire pipeline: image analysis, piece identification, board construction,
        and move calculation. One call gives you the answer.

        Args:
            image_source: Path to a chess board image file or URL.
            question: The chess question, including whose turn it is.
                Example: "It is black's turn. Provide the best move for black in algebraic notation."

        Returns:
            The move in algebraic notation (e.g., Qh7#) or analysis of the position.
        """
        return await asyncio.to_thread(_solve_chess_sync, image_source, question, config, api_key)

    yield FunctionInfo.from_fn(_solve_chess, description=_solve_chess.__doc__)
